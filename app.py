# -*- coding: utf-8 -*-
import os
import re
import uuid
import urllib3
import requests
import openpyxl
import openpyxl.styles.stylesheet
from openpyxl.styles import Font
import io
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, render_template, send_from_directory, send_file

# Salva il metodo originale ed applica la patch per evitare l'IndexError di openpyxl
original_expand = openpyxl.styles.stylesheet.Stylesheet._expand_named_style
def patched_expand(self, style):
    try:
        original_expand(self, style)
    except IndexError:
        pass
openpyxl.styles.stylesheet.Stylesheet._expand_named_style = patched_expand


# Disabilita gli avvisi per i certificati SSL non verificati
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'uploads')
app.config['PROCESSED_FOLDER'] = os.path.join(os.getcwd(), 'processed')

# Assicurati che le cartelle esistano
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)

# Registro globale dei job in background {job_id: {status, progress, messaggio, risultati, fileId, errore}}
jobs = {}

# Endpoint del catalogo GeoNetwork
URL_CATALOGO = "https://rsdi.regione.basilicata.it/geonetwork/srv/api/search/records/_search"

# Parole comuni GIS da ignorare o penalizzare nella ricerca specifica
PAROLE_GENERICHE_GIS = {
    'area', 'aree', 'zona', 'zone', 'mappa', 'mappe', 'carta', 'carte', 'punti', 'punto', 'linea', 'linee', 
    'poligono', 'poligoni', 'servizio', 'servizi', 'dati', 'dato', 'db', 'dbgt', 'basilicata', 'regione', 
    'provincia', 'provincie', 'comune', 'comuni', 'limiti', 'limite', 'confini', 'confine', 'wms', 'wfs', 
    'raster', 'vettoriale', 'shapefile', 'vista', 'volumi', 'nan', 'vigente', 'vigente1', 'vigente2', 'vigente3'
}

# Parole vuote in italiano da escludere (comprensive di articoli e preposizioni semplici e articolate)
PAROLE_VUOTE = {
    'il', 'la', 'lo', 'i', 'gli', 'le', 'un', 'una', 'uno', 'di', 'a', 'da', 'in', 'con', 'su', 'per', 'tra', 'fra',
    'del', 'dello', 'della', 'dei', 'degli', 'delle', 'al', 'allo', 'alla', 'ai', 'agli', 'alle', 
    'dal', 'dallo', 'dalla', 'dai', 'dagli', 'dalle', 'nel', 'nello', 'nella', 'nei', 'negli', 'nelle', 
    'col', 'coi', 'sul', 'sullo', 'sulla', 'sui', 'sugli', 'sulle', 'e', 'o', 'ed', 'ad', 'od', 'si', 'no', 'non'
}

# Sinonimi per mappare le discrepanze tra Excel e Catalogo
MAPPA_SINONIMI = {
    'incendi': ['fuoco', 'incendio'],
    'incendio': ['fuoco', 'incendi'],
    'fuoco': ['incendi', 'incendio'],
    'cave': ['coltivazione', 'mineraria'],
    'idro': ['idraulica', 'idraulico', 'idrografico', 'idrogeologico', 'idrografia', 'acque', 'idrico'],
    'idrografia': ['idrografico', 'idro', 'acque'],
    'idraulica': ['idro', 'idraulico', 'acque'],
    'idraulico': ['idro', 'idraulica', 'acque'],
    'nocciolo': ['corilicola', 'noccioli'],
    'corilicola': ['nocciolo', 'noccioli'],
    'coste': ['costa', 'coste', 'costiero', 'litorale', 'spiaggia', 'portuale'],
    'costa': ['coste', 'costiero', 'litorale', 'spiaggia', 'portuale'],
    'apistica': ['apistico', 'api'],
    'apistico': ['apistica', 'api'],
    'pedologica': ['pedologia', 'suolo'],
    'pedologia': ['pedologica', 'suolo'],
    'granulometrica': ['granulometria', 'suolo'],
    'tessitura': ['tessiture', 'suolo'],
    'carbonati': ['carbonato', 'suolo'],
    'reazione': ['ph', 'suolo'],
    'alluvioni': ['alluvione', 'idraulica', 'idraulico'],
    'alluvione': ['alluvioni', 'idraulica', 'idraulico'],
    'frane': ['frana', 'iffi', 'geologico', 'geologica', 'movimenti franosi'],
    'frana': ['frane', 'iffi', 'geologico', 'geologica', 'movimento franoso']
}

def pulisci_testo(testo):
    """Pulisce il testo sostituendo le parentesi con spazi ed eliminando caratteri non alfanumerici."""
    if not testo:
        return ""
    testo = str(testo).lower().strip()
    testo = testo.replace('(', ' ').replace(')', ' ')
    testo = re.sub(r'[^a-z0-9]', ' ', testo)
    testo = re.sub(r'\s+', ' ', testo).strip()
    return testo

def ottieni_token(testo):
    """Estrae i token significativi (escludendo le parole vuote)."""
    testo_pulito = pulisci_testo(testo)
    parole = testo_pulito.split()
    return [p for p in parole if p not in PAROLE_VUOTE and len(p) >= 2]

def ottieni_radice(parola):
    """Ritorna la radice (stem) di 5 caratteri per gestire le flessioni in italiano."""
    return parola[:5] if len(parola) >= 5 else parola

def verifica_valore_numerico_o_corto(parola):
    """Riconosce se una parola rappresenta un valore numerico, anno o scala (es: 2007, 10k)."""
    if parola.isdigit():
        return True
    if re.match(r'^\d+(k|m|cm|tr|l\d+|k\d+|g\d+)?$', parola):
        return True
    return False

def contiene_parola_intera(parola, testo):
    """Verifica se 'parola' appare come parola intera nel 'testo' (non come sottostringa)."""
    if not parola or not testo:
        return False
    return bool(re.search(r'\b' + re.escape(parola) + r'\b', testo))

def sono_incompatibili(layer_tokens, db_tokens, title_tokens):
    """Verifica se i token del layer/db e il titolo appartengono a categorie escludentesi a vicenda."""
    query_set = set(layer_tokens + db_tokens)
    titolo_set = set(title_tokens)
    
    # Gruppi mutuamente esclusivi
    incendi = {'incendio', 'incendi', 'fuoco', 'crdi'}
    frane = {'frane', 'frana', 'iffi', 'geologico', 'geologica', 'liguridi', 'suolo', 'suoli'}
    idraulica_acque = {'idraulica', 'idraulico', 'idro', 'idrografico', 'idrografia', 'acque', 'invasi', 'laghi', 'alluvione', 'alluvioni', 'idrico', 'fiume', 'sinni'}
    ortofoto = {'ortofoto', 'ortofotocarta'}
    apistica = {'apistica', 'apistico', 'api'}
    pedologia = {'pedologia', 'pedologica', 'granulometrica', 'tessitura', 'carbonati', 'reazione', 'suolo', 'suoli'}
    archeologia = {'archeologiche', 'archeologia'}
    
    categorie = [incendi, frane, idraulica_acque, ortofoto, apistica, pedologia, archeologia]
    
    cat_query = set()
    for i, cat in enumerate(categorie):
        if query_set & cat:
            cat_query.add(i)
            
    cat_titolo = set()
    for i, cat in enumerate(categorie):
        if titolo_set & cat:
            cat_titolo.add(i)
            
    if cat_query and cat_titolo and cat_query != cat_titolo:
        # Escludi valori numerici (anni, scale) dall'intersezione: "2007" in comune
        # non deve salvare un match tra categorie incompatibili
        intersezione_specifica = set()
        for t in (query_set & titolo_set):
            if t in PAROLE_GENERICHE_GIS or t in {'rischio', 'servizio'}:
                continue
            if verifica_valore_numerico_o_corto(t):
                continue
            intersezione_specifica.add(t)
        if not intersezione_specifica:
            return True
            
    return False

def valuta_corrispondenza(nome_layer, nome_db, titolo_record, uuid_record, identificatore_record, testo_completo_record):
    """Assegna un punteggio di corrispondenza tra una query dell'Excel e una scheda metadati."""
    if not titolo_record:
        return 0
        
    punteggio = 0
    
    layer_pulito = pulisci_testo(nome_layer)
    db_pulito = pulisci_testo(nome_db)
    titolo_pulito = pulisci_testo(titolo_record)
    
    token_layer = ottieni_token(nome_layer)
    token_db = ottieni_token(nome_db)
    token_titolo = ottieni_token(titolo_record)
    
    # Incompatibility check
    if sono_incompatibili(token_layer, token_db, token_titolo):
        return 0
        
    id_minuscolo = str(identificatore_record).lower()
    uuid_minuscolo = str(uuid_record).lower()
    testo_completo_minuscolo = str(testo_completo_record).lower()
    
    # Tokenizza anche il testo completo per matching a parola intera
    token_testo_completo = set(pulisci_testo(testo_completo_record).split())
    token_id = set(pulisci_testo(identificatore_record).split())
    token_uuid = set(pulisci_testo(uuid_record).split())
    
    # 1. Corrispondenza frasale esatta
    if layer_pulito and titolo_pulito == layer_pulito:
        punteggio += 300
    elif layer_pulito and contiene_parola_intera(layer_pulito, titolo_pulito):
        non_generici_in_layer = [t for t in token_layer if t not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(t)]
        if non_generici_in_layer:
            punteggio += 150
        else:
            punteggio += 40
            
    if db_pulito and db_pulito not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(db_pulito) and titolo_pulito == db_pulito:
        punteggio += 250
    elif db_pulito and db_pulito not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(db_pulito) and contiene_parola_intera(db_pulito, titolo_pulito):
        punteggio += 120

    if db_pulito and db_pulito not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(db_pulito):
        if db_pulito == id_minuscolo or contiene_parola_intera(db_pulito, id_minuscolo):
            punteggio += 200
        if contiene_parola_intera(db_pulito, uuid_minuscolo):
            punteggio += 180

    # 2. Corrispondenza basata su Token (solo parole esatte e sinonimi)
    corrispondenze_specifiche_non_numeriche = 0
    
    def controlla_corrispondenza_token(t):
        """Cerca un token nei token del titolo (match esatto di parola, non sottostringa)."""
        if t in token_titolo:
            return True, True
        if t in MAPPA_SINONIMI:
            for sinonimo in MAPPA_SINONIMI[t]:
                if sinonimo in token_titolo:
                    return True, False
        return False, False

    # Token del Layer
    specifiche_layer_trovate = 0
    generiche_layer_trovate = 0
    for t in token_layer:
        generico = t in PAROLE_GENERICHE_GIS
        numerico = verifica_valore_numerico_o_corto(t)
        
        trovato, esatto = controlla_corrispondenza_token(t)
        
        if trovato:
            valore = 40 if esatto else 25
            if generico:
                punteggio += 5
                generiche_layer_trovate += 1
            else:
                punteggio += valore
                specifiche_layer_trovate += 1
                if not numerico:
                    corrispondenze_specifiche_non_numeriche += 1
        elif t in token_id or t in token_uuid:
            # Match esatto a parola intera nell'identificatore o UUID
            if generico:
                punteggio += 2
            else:
                punteggio += 20
                specifiche_layer_trovate += 1
                if not numerico:
                    corrispondenze_specifiche_non_numeriche += 1
        elif t in token_testo_completo:
            # Match esatto a parola intera nel testo completo (punteggio ridotto)
            if not generico:
                punteggio += 3
                # NON contare come corrispondenza specifica: il testo completo è troppo ampio
                
    # Token del Database
    specifiche_db_trovate = 0
    for t in token_db:
        if t in PAROLE_GENERICHE_GIS:
            continue
        numerico = verifica_valore_numerico_o_corto(t)
        
        trovato, esatto = controlla_corrispondenza_token(t)
        if trovato:
            valore = 30 if esatto else 20
            punteggio += valore
            specifiche_db_trovate += 1
            if not numerico:
                corrispondenze_specifiche_non_numeriche += 1
        elif t in token_id or t in token_uuid:
            punteggio += 50
            specifiche_db_trovate += 1
            if not numerico:
                corrispondenze_specifiche_non_numeriche += 1
        elif t in token_testo_completo:
            # Match nel testo completo: punteggio minimo e NON conta come match specifico
            punteggio += 3

    # 3. Bonus per il completamento dei token specifici richiesti
    specifiche_layer_richieste = [t for t in token_layer if t not in PAROLE_GENERICHE_GIS]
    if specifiche_layer_richieste and specifiche_layer_trovate == len(specifiche_layer_richieste):
        punteggio += 50
        
    specifiche_db_richieste = [t for t in token_db if t not in PAROLE_GENERICHE_GIS]
    if specifiche_db_richieste and specifiche_db_trovate == len(specifiche_db_richieste):
        punteggio += 50

    # 4. Tie-breaker bonus: se il titolo inizia esattamente con il nome cercato
    if layer_pulito and titolo_pulito.startswith(layer_pulito):
        punteggio += 10
    elif db_pulito and titolo_pulito.startswith(db_pulito):
        punteggio += 10

    # REGOLA CRITICA 1: Se la richiesta contiene parole specifiche non numeriche,
    # almeno una deve coincidere nel TITOLO (non nel testo completo)
    specifiche_richieste_totali = [t for t in (token_layer + token_db) if t not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(t)]
    if specifiche_richieste_totali and corrispondenze_specifiche_non_numeriche == 0:
        return 0
    
    # REGOLA CRITICA 2: Richiedi una percentuale minima di corrispondenze specifiche
    # Per evitare match con 1 sola parola su molte
    totale_specifiche = len(specifiche_richieste_totali)
    if totale_specifiche >= 2 and corrispondenze_specifiche_non_numeriche < max(1, totale_specifiche // 2):
        # Se meno della metà delle parole specifiche corrisponde, penalizza pesantemente
        punteggio = punteggio // 3
    
    # REGOLA CRITICA 3: Se TUTTI i token specifici sono solo numerici (es. anni: 2007, 2008),
    # la query non ha contesto semantico sufficiente. Richiedi che almeno i token
    # generici/di contesto (es. "ortofoto") corrispondano nel titolo.
    token_tutti = [t for t in (token_layer + token_db) if t not in PAROLE_GENERICHE_GIS]
    token_non_numerici = [t for t in token_tutti if not verifica_valore_numerico_o_corto(t)]
    if token_tutti and not token_non_numerici:
        # Tutti i token specifici sono numerici - verifica il contesto generico
        token_contesto = [t for t in (token_layer + token_db) if t in PAROLE_GENERICHE_GIS]
        if token_contesto:
            contesto_trovato = any(t in token_titolo for t in token_contesto)
            if not contesto_trovato:
                # Il contesto generico (es. "ortofoto") non matcha il titolo → rifiuta
                return 0
        # Se non c'è nemmeno contesto generico (solo un anno), ridurre pesantemente
        # per evitare che "2007" da solo matchi qualsiasi record del 2007
        if not token_contesto:
            punteggio = punteggio // 4
    
    # REGOLA CRITICA 4: Se il nome_db contiene suffissi numerici (es. "apistica_1", "apistica_4"),
    # quei numeri devono essere presenti anche nel titolo del record.
    # Altrimenti "apistica_1" e "apistica_4" matcherebbero entrambi "Carta Apistica".
    if db_pulito:
        numeri_nel_db = [t for t in db_pulito.split() if t.isdigit()]
        if numeri_nel_db:
            # Controlla se almeno uno dei numeri del db è nel titolo
            parole_titolo = set(titolo_pulito.split())
            numeri_trovati = sum(1 for n in numeri_nel_db if n in parole_titolo)
            if numeri_trovati == 0:
                # Il suffisso numerico specifico non è nel titolo → match molto debole
                punteggio = punteggio // 4
        
    return punteggio

def prepara_catalogo(catalogo_raw):
    """Pre-processa tutte le schede del catalogo UNA SOLA VOLTA.
    
    Evita di ricalcolare tokenizzazione, pulizia testo e set di token
    per ogni confronto (da 2.4M operazioni ridondanti a 600).
    """
    risultato = []
    for scheda in catalogo_raw:
        source = scheda.get('_source', {})
        rto = source.get('resourceTitleObject')
        titolo = rto.get('langita', rto.get('default', '')) if isinstance(rto, dict) else ""
        if not titolo:
            continue  # Salta schede senza titolo (non matcheranno mai)
        
        uuid_sch = source.get('uuid', '')
        identificatore = source.get('resourceIdentifier', '')
        
        if isinstance(identificatore, list):
            parti_id = []
            for item in identificatore:
                if isinstance(item, dict):
                    parti_id.extend([str(v) for v in item.values() if v])
                else:
                    parti_id.append(str(item))
            identificatore = " ".join(parti_id)
        else:
            identificatore = str(identificatore)
            
        testo_completo = " ".join(source.get('anyText', [])) if isinstance(source.get('anyText'), list) else str(source.get('anyText', ''))
        
        titolo_pulito = pulisci_testo(titolo)
        token_titolo_list = ottieni_token(titolo)
        
        risultato.append({
            'scheda': scheda,
            'source': source,
            'titolo_pulito': titolo_pulito,
            'token_titolo_list': token_titolo_list,
            'token_titolo_set': set(token_titolo_list),   # O(1) lookup
            'id_minuscolo': identificatore.lower(),
            'uuid_minuscolo': uuid_sch.lower(),
            'token_id': set(pulisci_testo(identificatore).split()),
            'token_uuid': set(pulisci_testo(uuid_sch).split()),
            'token_testo_completo': set(pulisci_testo(testo_completo).split()),
        })
    return risultato


def valuta_corrispondenza_veloce(layer_pulito, db_pulito, token_layer, token_db, db_specifico, cat):
    """Versione ottimizzata di valuta_corrispondenza con dati pre-processati.
    
    Parametri pre-calcolati dalla riga:
        layer_pulito, db_pulito: testi puliti
        token_layer, token_db: token estratti
        db_specifico: bool - True se db_pulito è non-generico e non-numerico
    Parametri pre-calcolati dal catalogo:
        cat: dict con token_titolo_set, token_id, token_uuid, ecc.
    """
    punteggio = 0
    
    titolo_pulito = cat['titolo_pulito']
    token_titolo = cat['token_titolo_set']  # Set per lookup O(1)
    token_id = cat['token_id']
    token_uuid = cat['token_uuid']
    token_testo_completo = cat['token_testo_completo']
    id_minuscolo = cat['id_minuscolo']
    uuid_minuscolo = cat['uuid_minuscolo']
    
    # Incompatibility check
    if sono_incompatibili(token_layer, token_db, cat['token_titolo_list']):
        return 0
    
    # 1. Corrispondenza frasale esatta
    if layer_pulito and titolo_pulito == layer_pulito:
        punteggio += 300
    elif layer_pulito and contiene_parola_intera(layer_pulito, titolo_pulito):
        non_generici = [t for t in token_layer if t not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(t)]
        punteggio += 150 if non_generici else 40
            
    if db_specifico and titolo_pulito == db_pulito:
        punteggio += 250
    elif db_specifico and contiene_parola_intera(db_pulito, titolo_pulito):
        punteggio += 120

    if db_specifico:
        if db_pulito == id_minuscolo or contiene_parola_intera(db_pulito, id_minuscolo):
            punteggio += 200
        if contiene_parola_intera(db_pulito, uuid_minuscolo):
            punteggio += 180

    # 2. Token matching (con set O(1) invece di list O(n))
    corr_spec_non_num = 0
    
    specifiche_layer_trovate = 0
    for t in token_layer:
        generico = t in PAROLE_GENERICHE_GIS
        numerico = verifica_valore_numerico_o_corto(t)
        
        # Cerca nei token del titolo (O(1) con set)
        trovato_esatto = t in token_titolo
        trovato_sinonimo = False
        if not trovato_esatto and t in MAPPA_SINONIMI:
            for s in MAPPA_SINONIMI[t]:
                if s in token_titolo:
                    trovato_sinonimo = True
                    break
        
        if trovato_esatto or trovato_sinonimo:
            valore = 40 if trovato_esatto else 25
            if generico:
                punteggio += 5
            else:
                punteggio += valore
                specifiche_layer_trovate += 1
                if not numerico:
                    corr_spec_non_num += 1
        elif t in token_id or t in token_uuid:
            if generico:
                punteggio += 2
            else:
                punteggio += 20
                specifiche_layer_trovate += 1
                if not numerico:
                    corr_spec_non_num += 1
        elif not generico and t in token_testo_completo:
            punteggio += 3
                
    # Token del Database
    specifiche_db_trovate = 0
    for t in token_db:
        if t in PAROLE_GENERICHE_GIS:
            continue
        numerico = verifica_valore_numerico_o_corto(t)
        
        trovato_esatto = t in token_titolo
        trovato_sinonimo = False
        if not trovato_esatto and t in MAPPA_SINONIMI:
            for s in MAPPA_SINONIMI[t]:
                if s in token_titolo:
                    trovato_sinonimo = True
                    break
        
        if trovato_esatto or trovato_sinonimo:
            valore = 30 if trovato_esatto else 20
            punteggio += valore
            specifiche_db_trovate += 1
            if not numerico:
                corr_spec_non_num += 1
        elif t in token_id or t in token_uuid:
            punteggio += 50
            specifiche_db_trovate += 1
            if not numerico:
                corr_spec_non_num += 1
        elif t in token_testo_completo:
            punteggio += 3

    # 3. Bonus completamento
    spec_layer_rich = [t for t in token_layer if t not in PAROLE_GENERICHE_GIS]
    if spec_layer_rich and specifiche_layer_trovate == len(spec_layer_rich):
        punteggio += 50
    spec_db_rich = [t for t in token_db if t not in PAROLE_GENERICHE_GIS]
    if spec_db_rich and specifiche_db_trovate == len(spec_db_rich):
        punteggio += 50

    # 4. Tie-breaker
    if layer_pulito and titolo_pulito.startswith(layer_pulito):
        punteggio += 10
    elif db_pulito and titolo_pulito.startswith(db_pulito):
        punteggio += 10

    # Regole critiche
    spec_rich_tot = [t for t in (token_layer + token_db) if t not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(t)]
    if spec_rich_tot and corr_spec_non_num == 0:
        return 0
    
    tot_spec = len(spec_rich_tot)
    if tot_spec >= 2 and corr_spec_non_num < max(1, tot_spec // 2):
        punteggio = punteggio // 3
    
    token_tutti = [t for t in (token_layer + token_db) if t not in PAROLE_GENERICHE_GIS]
    token_non_num = [t for t in token_tutti if not verifica_valore_numerico_o_corto(t)]
    if token_tutti and not token_non_num:
        token_contesto = [t for t in (token_layer + token_db) if t in PAROLE_GENERICHE_GIS]
        if token_contesto:
            if not any(t in token_titolo for t in token_contesto):
                return 0
        if not token_contesto:
            punteggio = punteggio // 4
    
    if db_pulito:
        numeri_nel_db = [t for t in db_pulito.split() if t.isdigit()]
        if numeri_nel_db:
            parole_titolo = set(titolo_pulito.split())
            if sum(1 for n in numeri_nel_db if n in parole_titolo) == 0:
                punteggio = punteggio // 4
        
    return punteggio

def scarica_catalogo_rsdi():
    """Scarica l'elenco completo dei metadati dal catalogo RSDI Basilicata."""
    payload = {
        "size": 600,
        "query": {
            "match_all": {}
        }
    }
    intestazioni = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    try:
        risposta = requests.post(URL_CATALOGO, json=payload, headers=intestazioni, verify=False, timeout=20)
        if risposta.status_code == 200:
            dati = risposta.json()
            hits = dati.get('hits', {}).get('hits', [])
            # Esclude le schede template
            risultati = [h for h in hits if h.get('_source', {}).get('isTemplate') != 'y']
            return risultati
        else:
            print(f"Errore caricamento catalogo (Status {risposta.status_code})")
            return []
    except Exception as e:
        print(f"Eccezione durante il download del catalogo: {e}")
        return []

def formatta_data(valore_data):
    """Formatta la data in formato italiano DD-MM-YYYY."""
    if not valore_data:
        return ""
    val_str = str(valore_data).strip()
    match = re.match(r'^(\d{4})-(\d{2})-(\d{2})', val_str)
    if match:
        anno, mese, giorno = match.groups()
        return f"{giorno}-{mese}-{anno}"
    return val_str

def estrai_date_metadato(source_record):
    """Estrae le date di pubblicazione e revisione dal record GeoNetwork."""
    data_pub = ""
    data_rev = ""
    
    lista_pub = source_record.get('publicationDateForResource', [])
    lista_rev = source_record.get('revisionDateForResource', [])
    
    if lista_pub and isinstance(lista_pub, list):
        data_pub = lista_pub[0]
    if lista_rev and isinstance(lista_rev, list):
        data_rev = lista_rev[0]
        
    if not data_pub or not data_rev:
        date_risorsa = source_record.get('resourceDate', [])
        if date_risorsa and isinstance(date_risorsa, list):
            for rd in date_risorsa:
                if isinstance(rd, dict):
                    tipo = rd.get('type')
                    valore = rd.get('date')
                    if tipo == 'publication' and not data_pub:
                        data_pub = valore
                    elif tipo == 'revision' and not data_rev:
                        data_rev = valore
                        
    return formatta_data(data_pub), formatta_data(data_rev)

def _controlla_singolo_uid(uid):
    """Controlla un singolo UID per determinare il tipo di visualizzatore. Usato internamente dal pool."""
    url_nuovo_json = f"https://rsdi.regione.basilicata.it/crea-progetti/initFile/{uid}.json"
    url_nuovo_viewer = f"https://rsdi-view.regione.basilicata.it/#https://rsdi.regione.basilicata.it/crea-progetti/initFile/{uid}.json"
    url_standard = f"https://rsdi.regione.basilicata.it/viewGis/?project={uid}"
    
    try:
        risposta = requests.get(url_nuovo_json, verify=False, timeout=8)
        if risposta.status_code == 200:
            contenuto = risposta.text.strip()
            if contenuto and contenuto != '{}':
                return uid, ('NUOVO', url_nuovo_viewer)
    except Exception as e:
        print(f"Errore nel controllo visualizzatore per UID {uid}: {e}")
    
    return uid, ('STANDARD', url_standard)

def precarica_visualizzatori(lista_uid):
    """Controlla tutti gli UID in parallelo usando un ThreadPoolExecutor.
    
    Restituisce un dizionario {uid: ('NUOVO'|'STANDARD', link_visualizzatore)}.
    Con 10 worker paralleli, 298 UID vengono controllati in ~30 secondi invece di ~5 minuti.
    """
    cache = {}
    if not lista_uid:
        return cache
    
    print(f"Controllo visualizzatore per {len(lista_uid)} UID univoci in parallelo...")
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(_controlla_singolo_uid, uid): uid for uid in lista_uid}
        for future in as_completed(futures):
            try:
                uid, risultato = future.result()
                cache[uid] = risultato
            except Exception as e:
                uid = futures[future]
                print(f"Errore futuro per UID {uid}: {e}")
                cache[uid] = ('STANDARD', f"https://rsdi.regione.basilicata.it/viewGis/?project={uid}")
    
    nuovi = sum(1 for v in cache.values() if v[0] == 'NUOVO')
    print(f"Completato: {nuovi} NUOVO, {len(cache) - nuovi} STANDARD")
    return cache


@app.route('/')
def home():
    """Carica la pagina principale."""
    return render_template('index.html')


@app.route('/carica', methods=['POST'])
def carica_e_elabora():
    """Riceve il file Excel, lo salva e avvia l'elaborazione in background."""
    if 'file' not in request.files:
        return jsonify({"errore": "Nessun file inviato"}), 400
        
    file_excel = request.files['file']
    if file_excel.filename == '':
        return jsonify({"errore": "Nome del file vuoto"}), 400
        
    if not file_excel.filename.endswith('.xlsx'):
        return jsonify({"errore": "Il file deve essere in formato Excel (.xlsx)"}), 400

    # Genera identificativi univoci
    id_sessione = str(uuid.uuid4())
    nome_salvato = f"{id_sessione}_{file_excel.filename}"
    percorso_upload = os.path.join(app.config['UPLOAD_FOLDER'], nome_salvato)
    percorso_elaborato = os.path.join(app.config['PROCESSED_FOLDER'], nome_salvato)
    
    file_excel.save(percorso_upload)
    
    # Inizializza il job e avvia l'elaborazione in background
    job_id = id_sessione
    jobs[job_id] = {
        'status': 'processing',
        'progress': 0,
        'messaggio': 'Connessione al catalogo metadati RSDI...',
        'risultati': None,
        'fileId': None,
        'errore': None
    }
    
    thread = threading.Thread(
        target=elabora_in_background,
        args=(job_id, percorso_upload, percorso_elaborato, nome_salvato),
        daemon=True
    )
    thread.start()
    
    # Restituisce immediatamente il job_id (entro 1 secondo, nessun timeout)
    return jsonify({"jobId": job_id})


def elabora_in_background(job_id, percorso_upload, percorso_elaborato, nome_salvato):
    """Elabora il file Excel in un thread background, aggiornando il progresso nel registro jobs."""
    try:
        # Scarica il catalogo metadati
        jobs[job_id]['messaggio'] = 'Download catalogo metadati RSDI...'
        jobs[job_id]['progress'] = 5
        catalogo = scarica_catalogo_rsdi()
        if not catalogo:
            jobs[job_id]['status'] = 'error'
            jobs[job_id]['errore'] = 'Impossibile connettersi al catalogo RSDI Basilicata. Riprova più tardi.'
            return
        
        jobs[job_id]['progress'] = 10
        jobs[job_id]['messaggio'] = 'Apertura file Excel...'
        
        # Carica il workbook
        wb = openpyxl.load_workbook(percorso_upload)
        if not wb.sheetnames:
            jobs[job_id]['status'] = 'error'
            jobs[job_id]['errore'] = 'File Excel vuoto o non valido'
            return
            
        sheet = wb.active
        
        # Identifica gli indici delle colonne
        riga_intestazione = [str(cella.value).strip().lower() if cella.value is not None else "" for cella in sheet[1]]
        
        mappa_colonne = {}
        colonne_richieste = ['nome_completo', 'catalogo', 'data_pubblicazione', 'data_revisione', 'visualizzatore', 'uid', 'ufficio']
        for col_name in colonne_richieste:
            if col_name in riga_intestazione:
                mappa_colonne[col_name] = riga_intestazione.index(col_name) + 1
        
        # Per la colonna 'nome' (layer), prendi l'ULTIMA occorrenza
        indici_nome = [i + 1 for i, val in enumerate(riga_intestazione) if val == 'nome']
        if len(indici_nome) >= 2:
            mappa_colonne['nome_layer'] = indici_nome[-1]
        elif len(indici_nome) == 1:
            mappa_colonne['nome_layer'] = indici_nome[0]
                
        if 'nome_completo' not in mappa_colonne and 'nome_layer' not in mappa_colonne:
            jobs[job_id]['status'] = 'error'
            jobs[job_id]['errore'] = "Il file Excel deve contenere almeno le colonne 'nome_completo' o 'nome'."
            return
            
        # Crea colonne di output mancanti
        nuovo_indice_colonna = len(riga_intestazione) + 1
        for col_name in ['catalogo', 'data_pubblicazione', 'data_revisione', 'visualizzatore', 'ufficio']:
            if col_name not in mappa_colonne:
                sheet.cell(row=1, column=nuovo_indice_colonna, value=col_name)
                mappa_colonne[col_name] = nuovo_indice_colonna
                nuovo_indice_colonna += 1
                
        idx_nome_completo = mappa_colonne.get('nome_completo')
        idx_nome_layer = mappa_colonne.get('nome_layer')
        idx_catalogo = mappa_colonne.get('catalogo')
        idx_pub = mappa_colonne.get('data_pubblicazione')
        idx_rev = mappa_colonne.get('data_revisione')
        idx_visualizzatore = mappa_colonne.get('visualizzatore')
        idx_uid = mappa_colonne.get('uid')
        idx_ufficio = mappa_colonne.get('ufficio')
        
        # Pre-carica i visualizzatori in parallelo
        jobs[job_id]['messaggio'] = 'Controllo visualizzatori in parallelo...'
        jobs[job_id]['progress'] = 15
        
        uid_unici = set()
        for riga in range(2, sheet.max_row + 1):
            valore_uid = sheet.cell(row=riga, column=idx_uid).value if idx_uid else None
            if valore_uid:
                uid_unici.add(str(valore_uid).strip())
        cache_visualizzatore = precarica_visualizzatori(uid_unici)
        
        jobs[job_id]['progress'] = 35
        jobs[job_id]['messaggio'] = 'Pre-elaborazione catalogo...'
        
        # Pre-processa il catalogo UNA SOLA VOLTA (elimina 2.4M tokenizzazioni ridondanti)
        catalogo_pre = prepara_catalogo(catalogo)
        print(f"Catalogo pre-processato: {len(catalogo_pre)} schede valide")
        
        jobs[job_id]['progress'] = 40
        jobs[job_id]['messaggio'] = 'Analisi e confronto con il catalogo...'
        
        # Elaborazione righe
        cronologia_elaborazione = []
        soglia_matching = 80
        righe_totali = sheet.max_row - 1
        
        for idx_riga, riga in enumerate(range(2, sheet.max_row + 1)):
            valore_nome_completo = sheet.cell(row=riga, column=idx_nome_completo).value if idx_nome_completo else ""
            valore_nome_layer = sheet.cell(row=riga, column=idx_nome_layer).value if idx_nome_layer else ""
            valore_uid = sheet.cell(row=riga, column=idx_uid).value if idx_uid else ""
            
            if not valore_nome_completo and not valore_nome_layer:
                continue
                
            valore_nome_completo_str = str(valore_nome_completo).strip() if valore_nome_completo is not None else ""
            valore_nome_layer_str = str(valore_nome_layer).strip() if valore_nome_layer is not None else ""
            valore_uid_str = str(valore_uid).strip() if valore_uid is not None else ""
            
            # Pre-computa dati della riga UNA SOLA VOLTA (non 600 volte)
            layer_pulito = pulisci_testo(valore_nome_completo_str)
            db_pulito = pulisci_testo(valore_nome_layer_str)
            token_layer = ottieni_token(valore_nome_completo_str)
            token_db = ottieni_token(valore_nome_layer_str)
            db_specifico = bool(db_pulito and db_pulito not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(db_pulito))
            
            # Gestione namespace: pre-computa anche versione senza namespace
            db_pulito_senza_ns = ""
            token_db_senza_ns = []
            db_specifico_senza_ns = False
            if ':' in valore_nome_layer_str:
                nome_senza_ns = valore_nome_layer_str.split(':', 1)[1]
                db_pulito_senza_ns = pulisci_testo(nome_senza_ns)
                token_db_senza_ns = ottieni_token(nome_senza_ns)
                db_specifico_senza_ns = bool(db_pulito_senza_ns and db_pulito_senza_ns not in PAROLE_GENERICHE_GIS and not verifica_valore_numerico_o_corto(db_pulito_senza_ns))
            
            miglior_record = None
            punteggio_massimo = 0
            
            # Loop interno: usa dati pre-processati (nessuna tokenizzazione)
            for cat in catalogo_pre:
                score = valuta_corrispondenza_veloce(layer_pulito, db_pulito, token_layer, token_db, db_specifico, cat)
                
                if db_pulito_senza_ns:
                    score_ns = valuta_corrispondenza_veloce(layer_pulito, db_pulito_senza_ns, token_layer, token_db_senza_ns, db_specifico_senza_ns, cat)
                    if score_ns > score:
                        score = score_ns
                            
                if score > punteggio_massimo:
                    punteggio_massimo = score
                    miglior_record = cat['scheda']
            
            # Visualizzatore
            tipo_visualizzatore = ""
            link_visualizzatore = ""
            if valore_uid_str and valore_uid_str in cache_visualizzatore:
                tipo_visualizzatore, link_visualizzatore = cache_visualizzatore[valore_uid_str]
                    
            esito = "NO"
            titolo_trovato = ""
            link_trovato = "no"
            pub_data = ""
            rev_data = ""
            valore_ufficio = ""
            
            if punteggio_massimo >= soglia_matching:
                esito = "YES"
                source = miglior_record.get('_source', {})
                rto = source.get('resourceTitleObject')
                titolo_trovato = rto.get('langita', rto.get('default', '')) if isinstance(rto, dict) else ""
                uuid_trovato = source.get('uuid', '')
                link_trovato = f"https://rsdi.regione.basilicata.it/geonetwork/srv/ita/catalog.search#/metadata/{uuid_trovato}"
                pub_data, rev_data = estrai_date_metadato(source)
                
                org_list = source.get('orgForResource', source.get('OrgForResource', []))
                if org_list and isinstance(org_list, list):
                    primo_org = org_list[0]
                    if isinstance(primo_org, dict):
                        valore_ufficio = primo_org.get('default', primo_org.get('langita', ''))
                    else:
                        valore_ufficio = str(primo_org)
                elif isinstance(org_list, str):
                    valore_ufficio = org_list
                
                sheet.cell(row=riga, column=idx_catalogo, value=link_trovato)
                sheet.cell(row=riga, column=idx_pub, value=pub_data if pub_data else None)
                sheet.cell(row=riga, column=idx_rev, value=rev_data if rev_data else None)
                if idx_ufficio:
                    sheet.cell(row=riga, column=idx_ufficio, value=valore_ufficio if valore_ufficio else None)
            else:
                sheet.cell(row=riga, column=idx_catalogo, value="no")
                sheet.cell(row=riga, column=idx_pub, value=None)
                sheet.cell(row=riga, column=idx_rev, value=None)
                if idx_ufficio:
                    sheet.cell(row=riga, column=idx_ufficio, value=None)
            
            if idx_visualizzatore:
                sheet.cell(row=riga, column=idx_visualizzatore, value=tipo_visualizzatore if tipo_visualizzatore else None)
            if idx_uid and link_visualizzatore:
                sheet.cell(row=riga, column=idx_uid, value=link_visualizzatore)
                
            cronologia_elaborazione.append({
                "riga": riga,
                "nome_completo": valore_nome_completo_str,
                "nome": valore_nome_layer_str,
                "esito": esito,
                "titolo_metadato": titolo_trovato,
                "link": link_trovato,
                "pubblicazione": pub_data,
                "revisione": rev_data,
                "visualizzatore": tipo_visualizzatore,
                "ufficio": valore_ufficio,
                "punteggio": punteggio_massimo
            })
            
            # Aggiorna il progresso (40% -> 95%)
            if righe_totali > 0:
                avanzamento = 40 + int((idx_riga / righe_totali) * 55)
                jobs[job_id]['progress'] = min(avanzamento, 95)
            
        # Salva il file
        jobs[job_id]['messaggio'] = 'Salvataggio file Excel...'
        jobs[job_id]['progress'] = 96
        wb.save(percorso_elaborato)
        wb.close()
        
        if os.path.exists(percorso_upload):
            os.remove(percorso_upload)
            
        # Job completato
        jobs[job_id]['status'] = 'complete'
        jobs[job_id]['progress'] = 100
        jobs[job_id]['messaggio'] = 'Completato!'
        jobs[job_id]['fileId'] = nome_salvato
        jobs[job_id]['risultati'] = cronologia_elaborazione
        
    except Exception as e:
        if os.path.exists(percorso_upload):
            os.remove(percorso_upload)
        print(f"Errore durante l'elaborazione del file Excel: {e}")
        jobs[job_id]['status'] = 'error'
        jobs[job_id]['errore'] = f"Errore durante l'elaborazione: {str(e)}"


@app.route('/stato/<job_id>', methods=['GET'])
def stato_job(job_id):
    """Restituisce lo stato attuale di un job di elaborazione."""
    if job_id not in jobs:
        return jsonify({"errore": "Job non trovato"}), 404
    
    job = jobs[job_id]
    risposta = {
        'status': job['status'],
        'progress': job['progress'],
        'messaggio': job['messaggio']
    }
    
    if job['status'] == 'complete':
        risposta['fileId'] = job['fileId']
        risposta['risultati'] = job['risultati']
        # Pulizia: rimuovi il job dal registro dopo il recupero
        del jobs[job_id]
    elif job['status'] == 'error':
        risposta['errore'] = job['errore']
        del jobs[job_id]
    
    return jsonify(risposta)


@app.route('/scarica/<file_id>', methods=['GET'])
def scarica_file(file_id):
    """Permette al client di scaricare il file Excel elaborato."""
    # Sicurezza: convalida il nome del file
    if not re.match(r'^[a-f0-9\-]{36}_.+\.xlsx$', file_id):
        return jsonify({"errore": "ID file non valido"}), 400
        
    percorso_file = os.path.join(app.config['PROCESSED_FOLDER'], file_id)
    if not os.path.exists(percorso_file):
        return jsonify({"errore": "File non trovato o scaduto"}), 404
        
    # Rimuovi il prefisso UUID dal nome del file per il download dell'utente
    nome_originale = file_id[37:]
    
    ancorato = request.args.get('ancorato', 'false').lower() == 'true'
    
    if ancorato:
        try:
            wb = openpyxl.load_workbook(percorso_file)
            sheet = wb.active
            
            # Identifica gli indici delle colonne basandosi sull'intestazione (riga 1)
            riga_intestazione = [str(cella.value).strip().lower() if cella.value is not None else "" for cella in sheet[1]]
            idx_catalogo = None
            if 'catalogo' in riga_intestazione:
                idx_catalogo = riga_intestazione.index('catalogo') + 1
                
            if idx_catalogo:
                font_link = Font(color="0563C1", underline="single")
                for riga in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=riga, column=idx_catalogo)
                    val = str(cell.value or "").strip()
                    if val.startswith("http://") or val.startswith("https://"):
                        # Assegna l'hyperlink come stringa (preserva l'URL completo con #)
                        cell.hyperlink = val
                        cell.value = "sì"
                        cell.font = font_link
            
            # Salva in un buffer in memoria
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            wb.close()
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=nome_originale,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            print(f"Errore durante l'ancoraggio dei link nel download: {e}")
            # In caso di errore, fall back al file originale sul server
            pass
            
    return send_file(
        percorso_file,
        as_attachment=True,
        download_name=nome_originale,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)
